from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import json, os, io
from datetime import datetime
from contextlib import contextmanager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "promo-app-s3cr3t-key-2025")

DATA_FILE      = "data.json"
DATABASE_URL   = os.getenv("DATABASE_URL", "")
USE_DB         = bool(DATABASE_URL)
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "1029")

TEAMS = [
    "냉동BM1팀", "냉동BM2팀", "발효유팀", "Dairy제품팀",
    "커피 음료팀", "NC팀", "해외", "콘텐츠전략팀"
]

CATEGORIES         = ["신제품", "리뉴얼", "프로모션", "박람회"]
TASTING_CATEGORIES = ["신제품", "리뉴얼", "기타"]
TASTING_TEAMS      = [t for t in TEAMS if t not in ("해외", "콘텐츠전략팀")]

# ── DB 연결 (PostgreSQL) ──────────────────────────────────
if USE_DB:
    import psycopg2
    import psycopg2.extras

    def _db_url():
        url = DATABASE_URL
        return url.replace("postgres://", "postgresql://", 1) if url.startswith("postgres://") else url

    @contextmanager
    def db():
        conn = psycopg2.connect(_db_url())
        try:
            cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            yield cur
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def init_db():
        with db() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS entries (
                    id           TEXT PRIMARY KEY,
                    team         TEXT NOT NULL,
                    name         TEXT NOT NULL,
                    schedule     TEXT NOT NULL,
                    category     TEXT NOT NULL,
                    product      TEXT NOT NULL,
                    note         TEXT DEFAULT '',
                    submitted_at TEXT NOT NULL,
                    section      TEXT DEFAULT 'promo'
                )
            """)
            cur.execute("ALTER TABLE entries ADD COLUMN IF NOT EXISTS section TEXT DEFAULT 'promo'")
        print("[DB] table ready")

# ── 데이터 접근 ───────────────────────────────────────────
def load_data(team=None, section=None):
    if USE_DB:
        conditions, params = [], []
        if team:    conditions.append("team = %s");    params.append(team)
        if section: conditions.append("section = %s"); params.append(section)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        with db() as cur:
            cur.execute(f"SELECT * FROM entries {where} ORDER BY submitted_at DESC", params)
            return [dict(r) for r in cur.fetchall()]
    else:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if team:    data = [d for d in data if d.get("team") == team]
            if section: data = [d for d in data if d.get("section", "promo") == section]
            return data
        return []

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ── 라우트 ────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html",
        teams=TEAMS,
        categories=CATEGORIES,
        tasting_categories=TASTING_CATEGORIES,
        tasting_teams=TASTING_TEAMS)

@app.route("/admin")
def admin():
    return redirect(url_for("index"))

@app.route("/api/auth", methods=["POST"])
def auth():
    pw = request.json.get("password", "")
    if pw == ADMIN_PASSWORD:
        session["admin"] = True
        return jsonify({"ok": True})
    return jsonify({"ok": False}), 401

@app.route("/api/submit", methods=["POST"])
def submit():
    body    = request.json
    entries = body.get("entries", [])
    section = body.get("section", "promo")
    if not entries:
        return jsonify({"success": False, "message": "데이터가 없습니다."}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if USE_DB:
        with db() as cur:
            for i, entry in enumerate(entries):
                eid = f"{now}_{i}_{os.urandom(3).hex()}"
                cur.execute("""
                    INSERT INTO entries
                        (id, team, name, schedule, category, product, note, submitted_at, section)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (eid, entry["team"], entry["name"], entry["schedule"],
                      entry["category"], entry["product"], entry.get("note", ""), now, section))
    else:
        data = load_data()
        for i, entry in enumerate(entries):
            entry.update({"submitted_at": now, "id": f"{now}_{len(data)+i}", "section": section})
            data.append(entry)
        save_data(data)

    return jsonify({"success": True, "message": f"{len(entries)}건 제출 완료!"})

@app.route("/api/data")
def get_data():
    return jsonify(load_data(team=request.args.get("team"), section=request.args.get("section")))

@app.route("/api/delete/<entry_id>", methods=["DELETE"])
def delete_entry(entry_id):
    if not session.get("admin"):
        return jsonify({"success": False, "message": "권한이 없습니다."}), 403
    if USE_DB:
        with db() as cur:
            cur.execute("DELETE FROM entries WHERE id = %s", (entry_id,))
    else:
        save_data([d for d in load_data() if d.get("id") != entry_id])
    return jsonify({"success": True})

@app.route("/api/clear_team", methods=["POST"])
def clear_team():
    if not session.get("admin"):
        return jsonify({"success": False, "message": "권한이 없습니다."}), 403
    team = request.json.get("team")
    if USE_DB:
        with db() as cur:
            cur.execute("DELETE FROM entries WHERE team = %s", (team,))
    else:
        save_data([d for d in load_data() if d.get("team") != team])
    return jsonify({"success": True})

# ── 엑셀 내보내기 ─────────────────────────────────────────
@app.route("/export")
def export():
    section = request.args.get("section", "promo")
    data    = load_data(section=section)

    wb          = openpyxl.Workbook()
    header_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    body_font   = Font(name="맑은 고딕", size=10)
    navy        = PatternFill("solid", fgColor="1F3864")
    white       = PatternFill("solid", fgColor="FFFFFF")
    stripe      = PatternFill("solid", fgColor="F2F7FB")
    thin        = Side(style="thin", color="B0C4D8")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    COLS   = ["일정", "담당팀", "담당자", "구분", "제품명", "비고"]
    WIDTHS = [14, 14, 10, 10, 32, 40]
    title  = "신제품 체험단 일정 취합본" if section == "tasting" else "제품 출시 및 프로모션 일정 취합본"

    def make_sheet(ws, rows):
        ws.merge_cells("A1:F1")
        ws["A1"] = title
        ws["A1"].font      = Font(name="맑은 고딕", bold=True, size=14, color="1F3864")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws["A1"].fill      = white

        ws.merge_cells("A2:F2")
        ws["A2"] = f"작성일자: {datetime.now().strftime('%Y.%m.%d')}    작성팀: 홍보담당 콘텐츠전략팀"
        ws["A2"].font      = Font(name="맑은 고딕", size=9, color="595959")
        ws["A2"].alignment = left

        ws.merge_cells("A3:F3")
        ws["A3"] = "※ 아래는 당월 및 향후 2개월 일정으로, 익월 이후 일정은 변동 가능성에 유의"
        ws["A3"].font      = Font(name="맑은 고딕", size=9, italic=True, color="C00000")
        ws["A3"].alignment = left

        ws.merge_cells("A4:F4"); ws["A4"].fill = white

        for c_idx, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
            cell           = ws.cell(row=5, column=c_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = navy
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = width
        ws.row_dimensions[5].height = 20

        for r_idx, row in enumerate(rows, 6):
            fill = stripe if r_idx % 2 == 0 else white
            ws.row_dimensions[r_idx].height = 18
            for c_idx, key in enumerate(["schedule","team","name","category","product","note"], 1):
                cell           = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
                cell.font      = body_font
                cell.fill      = fill
                cell.alignment = center if c_idx <= 4 else left
                cell.border    = border

        ws.freeze_panes = "A6"
        for i in range(1, 5):
            ws.row_dimensions[i].height = 18

    def sort_key(d):
        s = d.get("schedule", "")
        try:
            dt = datetime.strptime(s, "%Y-%m-%d")
            return (dt.year, dt.month, dt.day, d.get("team",""), d.get("name",""))
        except ValueError:
            pass
        try:
            dt = datetime.strptime(s, "%Y-%m")
            return (dt.year, dt.month, 0, d.get("team",""), d.get("name",""))
        except ValueError:
            pass
        MONTH_ORDER  = {m: i for i,m in enumerate(["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"])}
        PERIOD_ORDER = {"초":0,"중":1,"말":2,"(미정)":3}
        for m, mi in MONTH_ORDER.items():
            if s.startswith(m):
                rest = s[len(m):]
                pi   = next((v for k,v in PERIOD_ORDER.items() if k in rest), 99)
                return (2024, mi+1, pi*10+1, d.get("team",""), d.get("name",""))
        return (9999, 12, 31, d.get("team",""), d.get("name",""))

    sorted_data  = sorted(data, key=sort_key)
    ws_all       = wb.active
    ws_all.title = "취합"
    make_sheet(ws_all, sorted_data)

    for team in TEAMS:
        team_data = [d for d in sorted_data if d.get("team") == team]
        if not team_data: continue
        ws = wb.create_sheet(title=team[:15])
        make_sheet(ws, team_data)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    prefix   = "체험단일정" if section == "tasting" else "프로모션일정"
    filename = f"{prefix}_취합본_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── 시작 ─────────────────────────────────────────────────
if USE_DB:
    with app.app_context():
        init_db()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
